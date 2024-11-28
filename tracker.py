import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

def get_total_pages(soup):
    pagination = soup.find('ul', class_='page-numbers')
    if pagination:
        page_numbers = [int(link.text) for link in pagination.find_all('a', class_='page-numbers') 
                       if link.text.isdigit()]
        return max(page_numbers) if page_numbers else 1
    return 1

def clean_price(price_text):
    if not price_text or price_text == 'N/A':
        return 'بدون قیمت'
    return price_text.strip()

def scrape_page(url, headers):
    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        
        products = soup.find_all(['li', 'div'], class_=['product', 'product-item', 'product-grid-item'])
        
        product_list = []
        for product in products:
            title_tag = (
                product.find('h3', class_=['wd-entities-title', 'product-title']) or
                product.find('h2', class_=['wd-entities-title', 'product-title']) or
                product.find('a', class_=['product-title'])
            )
            
            price_tag = (
                product.find('span', class_='woocommerce-Price-amount') or
                product.find('span', class_='price') or
                product.find('div', class_='price')
            )
            
            title = title_tag.get_text(strip=True) if title_tag else 'N/A'
            price = clean_price(price_tag.get_text(strip=True) if price_tag else None)
            current_time = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
            
            product_list.append([title, price, current_time])
            
        return product_list
    except Exception as e:
        print(f"خطا در استخراج صفحه {url}: {e}")
        return []

def update_product_list(existing_df, new_products):
    columns = ['نام محصول', 'قیمت', 'تاریخ بروزرسانی']
    new_df = pd.DataFrame(new_products, columns=columns)
    
    if existing_df is not None and not existing_df.empty:
        try:
            # تبدیل DataFrame‌ها به دیکشنری برای مقایسه سریع‌تر
            existing_dict = existing_df.set_index('نام محصول').to_dict('index')
            updated_products = []
            
            for _, row in new_df.iterrows():
                name = row['نام محصول']
                if name in existing_dict:
                    # به‌روزرسانی قیمت و تاریخ برای محصولات موجود
                    updated_product = [
                        name,
                        row['قیمت'],
                        row['تاریخ بروزرسانی']
                    ]
                else:
                    # اضافه کردن محصول جدید
                    updated_product = [
                        name,
                        row['قیمت'],
                        row['تاریخ بروزرسانی']
                    ]
                updated_products.append(updated_product)
            
            return pd.DataFrame(updated_products, columns=columns)
        except Exception as e:
            print(f"خطا در به‌روزرسانی لیست محصولات: {e}")
            return new_df
    
    return new_df

def format_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active
    
    # تعریف رنگ‌ها
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    alternate_fill = PatternFill(start_color='E9EDF4', end_color='E9EDF4', fill_type='solid')
    
    # تنظیم راست به چپ برای همه سلول‌ها و اعمال رنگ‌بندی یک در میان
    row_count = ws.max_row
    col_count = ws.max_column
    
    for row in range(1, row_count + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # رنگ‌بندی هدر
            if row == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # رنگ‌بندی ردیف‌های زوج
            elif row % 2 == 0:
                cell.fill = alternate_fill
            
    # تنظیم عرض ستون‌ها
    ws.column_dimensions['A'].width = 40  # نام محصول
    ws.column_dimensions['B'].width = 20  # قیمت
    ws.column_dimensions['C'].width = 20  # تاریخ بروزرسانی
    
    # اعمال راست به چپ برای کل sheet
    ws.sheet_view.rightToLeft = True
    
    # اعمال فریز برای ردیف هدر
    ws.freeze_panes = 'A2'
    
    wb.save(filename)

def scrape_saterco_shop():
    base_url = 'https://www.yoursite.com//shop/page/{}/'
    first_page_url = 'https://yoursite.com/shop/'
    excel_file = 'all_products_list.xlsx'
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    try:
        existing_df = None
        if os.path.exists(excel_file):
            try:
                existing_df = pd.read_excel(excel_file)
                print(f"فایل قبلی با {len(existing_df)} محصول یافت شد.")
            except Exception as e:
                print(f"خطا در خواندن فایل اکسل: {e}")
        
        response = requests.get(first_page_url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        total_pages = get_total_pages(soup)
        
        print(f"تعداد کل صفحات: {total_pages}")
        
        all_products = []
        
        print("در حال استخراج صفحه 1...")
        all_products.extend(scrape_page(first_page_url, headers))
        
        for page in range(2, total_pages + 1):
            print(f"در حال استخراج صفحه {page}...")
            page_url = base_url.format(page)
            products = scrape_page(page_url, headers)
            all_products.extend(products)
            time.sleep(1)
        
        if all_products:
            updated_df = update_product_list(existing_df, all_products)
            
            print("\nذخیره در فایل اکسل...")
            updated_df.to_excel(excel_file, index=False)
            
            print("قالب‌بندی فایل اکسل...")
            format_excel(excel_file)
            
            print(f"\nتعداد کل محصولات: {len(updated_df)}")
            if existing_df is not None:
                new_products = len(updated_df) - len(existing_df)
                print(f"تعداد محصولات جدید اضافه شده: {new_products}")
            print("لیست محصولات با موفقیت به‌روزرسانی شد.")
            
            return updated_df
        else:
            print("هیچ محصولی یافت نشد!")
            return None
            
    except Exception as e:
        print(f"خطای غیرمنتظره: {e}")
        return None

if __name__ == "__main__":
    df = scrape_saterco_shop()
    if df is not None:
        print("\nنمونه داده‌های به‌روزرسانی شده:")
        print(df.head())
